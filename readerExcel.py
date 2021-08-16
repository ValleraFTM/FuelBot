
def readExcel(filename):
    """функция получает тип ТС и загружает соответствующий файл эксель
    и возвращает список со списком в формате [ТС, описание, нормы ГСМ]"""

    import openpyxl

    wb = openpyxl.load_workbook('./baza_norm/' + filename + '.xlsx')
    subs = []
    sheets = wb.sheetnames
    for sheet in sheets:
        subs.append(sheet)
    print(subs)
    sheet = wb.active
    string_xlsx = sheet.max_row
    print(string_xlsx)

    car_list = []
    for i in range(1, string_xlsx + 1):
        car = []
        for row in sheet[i]:
            if row.value == None:
                break
            car.append(str(row.value))
        car_list.append(car)
    print(car_list)
    return car_list
